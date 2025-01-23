import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from tkinter import Tk, filedialog, messagebox
import os

# Função para abrir a janela de seleção de arquivo
def selecionar_arquivo():
    Tk().withdraw()  # Ocultar a janela principal do Tkinter
    caminho_arquivo = filedialog.askopenfilename(title="Selecione o arquivo Excel", filetypes=[("Arquivos Excel", "*.xlsx")])
    return caminho_arquivo

# Função para mostrar uma mensagem para o usuário e registrar no arquivo de log
def mostrar_mensagem(titulo, mensagem, arquivo_log):
    messagebox.showwarning(titulo, mensagem)
    arquivo_log.write(f"{titulo}: {mensagem}\n")

# Definindo o valor de comprimento_maximo (pode ser alterado conforme necessário)
comprimento_maximo = 5.9

# Carregar o arquivo selecionado
caminho_arquivo = selecionar_arquivo()
if not caminho_arquivo:
    print("Nenhum arquivo selecionado.")
    exit()

# Carregar os dados da planilha Excel
dataset = pd.read_excel(caminho_arquivo, sheet_name="Lista", skiprows=1)

# Filtrar os dados relevantes
df = dataset[['Material', 'Quant.', 'Dimensão', 'Unid. Medida', 'TAG DO CONJUNTO']]

# Criação do arquivo de log
diretorio = os.path.dirname(caminho_arquivo)
nome_log = f"{os.path.splitext(os.path.basename(caminho_arquivo))[0]}_mensagens.txt"
caminho_log = os.path.join(diretorio, nome_log)

# Abrir o arquivo de log para escrita
with open(caminho_log, "w") as arquivo_log:
    # Verificação de itens com comprimento maior que o limite
    df_maior_que_maximo = df[df['Dimensão'] > comprimento_maximo]
    if not df_maior_que_maximo.empty:
        for _, row in df_maior_que_maximo.iterrows():
            mensagem = f"Item: {row['Material']} com TAG: {row['TAG DO CONJUNTO']} tem comprimento de {row['Dimensão']}m, maior que o limite de {comprimento_maximo}m. Linha: {row.name + 3}"
            mostrar_mensagem("Erro: Item com Tamanho Maior que o Limite", mensagem, arquivo_log)

    # Verificação de campos nulos ou vazios na coluna 'Material', apenas nas linhas com 'Quant.' preenchido
    df_nulos_material = df[df['Quant.'].notnull() & (df['Material'].isnull() | (df['Material'].str.strip() == ''))]

    if not df_nulos_material.empty:
        for _, row in df_nulos_material.iterrows():
            mensagem = f"Campo 'Material' nulo ou vazio encontrado para TAG: {row['TAG DO CONJUNTO']}. Linha: {row.name + 3}"
            mostrar_mensagem("Erro: Campo Nulo ou Vazio no Material", mensagem, arquivo_log)

    # Filtrar apenas os itens com unidade de medida "m"
    df_metros = df[df['Unid. Medida'] == 'm']

    # Converter a coluna 'Dimensão' para numérica
    df_metros['Dimensão'] = pd.to_numeric(df_metros['Dimensão'], errors='coerce')

    # Agrupar itens por material e fazer a alocação de itens nas peças
    def agrupar_itens_por_material(df, comprimento_maximo):
        resultado = []
        for material, grupo in df.groupby('Material'):
            grupo = grupo.sort_values(by='Dimensão', ascending=False)
            pecas = []
            itens = []
            for _, item in grupo.iterrows():
                quantidade = int(item['Quant.'])
                for _ in range(quantidade):
                    itens.append(item)

            while itens:
                item = itens.pop(0)
                alocado = False
                for peca in pecas:
                    comprimento_restante = comprimento_maximo - sum(i['Dimensão'] for i in peca)
                    if item['Dimensão'] <= comprimento_restante:
                        peca.append(item)
                        alocado = True
                        break
                if not alocado:
                    pecas.append([item])
            resultado.append((material, pecas))
        return resultado

    # Consolidar os itens dentro de uma peça para exibição
    def consolidar_peca(peca):
        consolidado = {}
        for item in peca:
            chave = (item['Material'], item['Dimensão'], item['TAG DO CONJUNTO'])
            if chave in consolidado:
                consolidado[chave] += 1
            else:
                consolidado[chave] = 1
        return consolidado

    # Função para calcular o total arredondado com base no número de peças
    def calcular_total_arredondado_por_pecas(pecas):
        total_arredondado = len(pecas) * 6  # Cada peça é arredondada para 6 metros
        return total_arredondado

    # Organizar os itens em peças
    pecas_resultado = agrupar_itens_por_material(df_metros, comprimento_maximo)

    # Carregar o arquivo original para adicionar a nova aba
    wb_original = load_workbook(caminho_arquivo)

    # Criar uma nova aba chamada "Plano de Corte"
    if "Plano de Corte" not in wb_original.sheetnames:
        wb_original.create_sheet("Plano de Corte")
    sheet = wb_original["Plano de Corte"]

    # Configurar estilo
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal='center', vertical='center')
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Adicionar título
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    sheet.cell(row=1, column=1, value="Plano de Corte").font = title_font
    sheet.cell(row=1, column=1).alignment = center_alignment

    current_row = 2

    # Processar os dados e adicionar na aba "Plano de Corte"
    for material, pecas in pecas_resultado:
        sheet.cell(row=current_row, column=1, value="Material").font = header_font
        sheet.cell(row=current_row, column=2, value=material).font = Font(bold=True)
        current_row += 1

        for i, peca in enumerate(pecas, 1):
            comprimento_total = sum(item['Dimensão'] for item in peca)
            sheet.cell(row=current_row, column=1, value=f"Peça {i}").font = header_font
            sheet.cell(row=current_row, column=2, value=f"Comprimento Total: {comprimento_total:.2f} m").font = Font(bold=True)
            current_row += 1

            consolidado = consolidar_peca(peca)
            for (material, dimensao, tag), quantidade in consolidado.items():
                cell_dimensao = sheet.cell(row=current_row, column=2, value=f"{dimensao * 1000:.0f} mm")
                if dimensao > comprimento_maximo:
                    cell_dimensao.fill = red_fill  # Destacar em vermelho
                sheet.cell(row=current_row, column=1, value="Comprimento")
                sheet.cell(row=current_row, column=3, value=quantidade)
                sheet.cell(row=current_row, column=4, value=tag)
                current_row += 1

    # Criar uma nova aba chamada "Agrupamento"
    if "Agrupamento" not in wb_original.sheetnames:
        wb_original.create_sheet("Agrupamento")
    agrupamento_sheet = wb_original["Agrupamento"]

    agrupamento_title_font = Font(bold=True, size=14)
    agrupamento_header_font = Font(bold=True, size=12)

    agrupamento_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    agrupamento_sheet.cell(row=1, column=1, value="Agrupamento").font = agrupamento_title_font
    agrupamento_sheet.cell(row=1, column=1).alignment = center_alignment

    # Agrupar dados por Material e Dimensão e ordenar por Dimensão em ordem decrescente
    agrupamento_df = df_metros.groupby(['Material', 'Dimensão']).agg({'Quant.': 'sum'}).reset_index()
    agrupamento_df = agrupamento_df.sort_values(by='Dimensão', ascending=False)

    current_row = 2
    for material, material_group in agrupamento_df.groupby('Material'):
        agrupamento_sheet.cell(row=current_row, column=1, value="Material").font = agrupamento_header_font
        agrupamento_sheet.cell(row=current_row, column=2, value=material).font = Font(bold=True)
        current_row += 1

        dimensao_quantidade = []
        for _, row in material_group.iterrows():
            agrupamento_sheet.cell(row=current_row, column=1, value="Dimensão (mm)")
            agrupamento_sheet.cell(row=current_row, column=2, value=f"{row['Dimensão'] * 1000:.0f} mm")  # Convertendo para mm
            agrupamento_sheet.cell(row=current_row, column=3, value=row['Quant.'])
            current_row += 1

        # Calcular o total arredondado para o número de peças
        pecas = [peca for _, pecas in pecas_resultado if _ == material for peca in pecas]
        total_arredondado = calcular_total_arredondado_por_pecas(pecas)

        # Adicionar o total arredondado na última linha do agrupamento
        agrupamento_sheet.cell(row=current_row, column=1, value="Total para compra:")
        agrupamento_sheet.cell(row=current_row, column=2, value=f"{total_arredondado:.1f} m")  # Já em metros
        current_row += 1

    # Salvar o arquivo com um novo nome
    nome_otimizado = f"{os.path.splitext(os.path.basename(caminho_arquivo))[0]}_otimizada.xlsx"
    caminho_otimizado = os.path.join(diretorio, nome_otimizado)

    wb_original.save(caminho_otimizado)
    print(f"Arquivo salvo como '{caminho_otimizado}'")

    # Confirmar se o arquivo de log foi criado
    if os.path.exists(caminho_log):
        print(f"Arquivo de mensagens salvo como '{caminho_log}'")
        messagebox.showinfo("Arquivo de Log", f"O arquivo de log foi gerado: {caminho_log}")
